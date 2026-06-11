#!/usr/bin/env python3
"""
POA&M generator / validator for Trivy vulnerability scan results.

Modes (auto-detected from the presence of --poam-file):
  initial   - No existing POA&M state. Every finding becomes a new Open item
              with its own POA&M tracking ID.
  validate  - Existing POA&M state found. Diff current scan against tracked items:
                * still present            -> stays Open, last_seen updated
                * no longer present        -> moved to Closed (closed_date set)
                * newly seen               -> new Open item, new POA&M ID
                * Closed but seen again    -> Reopened (keeps original POA&M ID)

Gate:
  If any NEW or REOPENED item this run has severity >= --fail-on (and is not
  covered by an unexpired waiver), the script exits with code 2 AFTER writing
  all outputs, so the workflow can still upload artifacts with `if: always()`.

Exit codes:
  0 - gate passed
  1 - script / usage error
  2 - gate failed (new or reopened findings at/above the fail-on threshold)

Identity:
  A finding is keyed on (CVE ID, package name). Version churn on the same
  package does not create duplicate POA&M items; installed/fixed versions are
  refreshed on each scan.
"""

import argparse
import json
import os
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("ERROR: openpyxl is required (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

SCHEMA_VERSION = "1.0"
SEVERITY_RANK = {"CRITICAL": 4, "HIGH": 3, "MEDIUM": 2, "LOW": 1, "UNKNOWN": 0}
SEVERITY_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNKNOWN"]
DEFAULT_SLA = {"CRITICAL": 15, "HIGH": 30, "MEDIUM": 90, "LOW": 180, "UNKNOWN": 180}

EXIT_OK = 0
EXIT_ERROR = 1
EXIT_GATE_FAILED = 2

SEVERITY_FILL = {
    "CRITICAL": "C0392B",
    "HIGH": "E67E22",
    "MEDIUM": "F1C40F",
    "LOW": "27AE60",
    "UNKNOWN": "95A5A6",
}


def today_iso() -> str:
    return date.today().isoformat()


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def norm_severity(value: str) -> str:
    value = (value or "UNKNOWN").upper()
    return value if value in SEVERITY_RANK else "UNKNOWN"


def parse_sla(spec: str) -> dict:
    sla = dict(DEFAULT_SLA)
    for part in (spec or "").split(","):
        part = part.strip()
        if not part:
            continue
        if "=" not in part:
            raise ValueError(f"Bad SLA entry '{part}', expected SEVERITY=DAYS")
        sev, days = part.split("=", 1)
        sla[norm_severity(sev)] = int(days)
    return sla


# --------------------------------------------------------------------------
# Trivy parsing
# --------------------------------------------------------------------------

def load_trivy(paths: list) -> dict:
    """Return {key: finding} where key = 'CVE::package'."""
    findings = {}
    for p in paths:
        path = Path(p)
        if not path.is_file():
            raise FileNotFoundError(f"Trivy results file not found: {p}")
        data = json.loads(path.read_text())
        artifact = data.get("ArtifactName", "")
        for result in data.get("Results") or []:
            target = result.get("Target", "")
            for v in result.get("Vulnerabilities") or []:
                cve = v.get("VulnerabilityID")
                if not cve:
                    continue
                pkg = v.get("PkgName", "")
                key = f"{cve}::{pkg}"
                sev = norm_severity(v.get("Severity"))
                existing = findings.get(key)
                if existing is None:
                    findings[key] = {
                        "cve_id": cve,
                        "package": pkg,
                        "installed_version": v.get("InstalledVersion", ""),
                        "fixed_version": v.get("FixedVersion", ""),
                        "severity": sev,
                        "title": (v.get("Title") or "").strip(),
                        "primary_url": v.get("PrimaryURL", ""),
                        "targets": [target] if target else [],
                        "artifact": artifact,
                    }
                else:
                    if target and target not in existing["targets"]:
                        existing["targets"].append(target)
                    if SEVERITY_RANK[sev] > SEVERITY_RANK[existing["severity"]]:
                        existing["severity"] = sev
    return findings


# --------------------------------------------------------------------------
# Waivers
# --------------------------------------------------------------------------

def load_waivers(path: str) -> list:
    if not path:
        return []
    p = Path(path)
    if not p.is_file():
        print(f"WARNING: waivers file '{path}' not found, ignoring", file=sys.stderr)
        return []
    waivers = json.loads(p.read_text())
    if not isinstance(waivers, list):
        raise ValueError("Waivers file must be a JSON list")
    return waivers


def match_waiver(item: dict, waivers: list):
    """Return the first unexpired waiver matching this item, else None."""
    for w in waivers:
        if w.get("cve_id") != item["cve_id"]:
            continue
        if w.get("package") and w["package"] != item["package"]:
            continue
        expires = w.get("expires", "")
        if expires and expires < today_iso():
            continue  # expired waiver no longer protects
        return w
    return None


# --------------------------------------------------------------------------
# POA&M state
# --------------------------------------------------------------------------

def new_poam(system_name: str, image_ref: str) -> dict:
    return {
        "metadata": {
            "schema_version": SCHEMA_VERSION,
            "system_name": system_name,
            "image_ref": image_ref,
            "created": now_utc(),
            "last_updated": now_utc(),
            "scan_count": 0,
            "next_poam_seq": 1,
        },
        "items": [],
    }


def next_id(meta: dict) -> str:
    seq = meta["next_poam_seq"]
    meta["next_poam_seq"] = seq + 1
    return f"POAM-{seq:04d}"


def make_item(poam_id: str, finding: dict, sla: dict, image_ref: str) -> dict:
    sev = finding["severity"]
    detected = today_iso()
    due = (date.today() + timedelta(days=sla.get(sev, 180))).isoformat()
    return {
        "poam_id": poam_id,
        "status": "Open",
        "cve_id": finding["cve_id"],
        "package": finding["package"],
        "installed_version": finding["installed_version"],
        "fixed_version": finding["fixed_version"],
        "severity": sev,
        "title": finding["title"],
        "primary_url": finding["primary_url"],
        "targets": finding["targets"],
        "image": image_ref or finding.get("artifact", ""),
        "first_detected": detected,
        "last_seen": detected,
        "scheduled_completion": due,
        "closed_date": None,
        "reopened_date": None,
        "history": [{"date": detected, "event": "detected"}],
    }


def item_key(item: dict) -> str:
    return f"{item['cve_id']}::{item['package']}"


def process(poam: dict, findings: dict, sla: dict, image_ref: str, waivers: list) -> dict:
    """Mutates poam in place. Returns run statistics."""
    meta = poam["metadata"]
    meta["scan_count"] += 1
    meta["last_updated"] = now_utc()
    if image_ref:
        meta["image_ref"] = image_ref

    stats = {"new": [], "closed": [], "reopened": [], "still_open": []}
    remaining = dict(findings)
    t = today_iso()

    for item in poam["items"]:
        key = item_key(item)
        current = remaining.pop(key, None)
        if current is not None:
            # Refresh volatile fields from the latest scan
            item["installed_version"] = current["installed_version"]
            item["fixed_version"] = current["fixed_version"]
            item["targets"] = current["targets"]
            if SEVERITY_RANK[current["severity"]] > SEVERITY_RANK[item["severity"]]:
                item["severity"] = current["severity"]
            item["last_seen"] = t
            if item["status"] == "Closed":
                item["status"] = "Open"
                item["reopened_date"] = t
                item["closed_date"] = None
                item["scheduled_completion"] = (
                    date.today() + timedelta(days=sla.get(item["severity"], 180))
                ).isoformat()
                item["history"].append({"date": t, "event": "reopened"})
                stats["reopened"].append(item)
            else:
                stats["still_open"].append(item)
        else:
            if item["status"] == "Open":
                item["status"] = "Closed"
                item["closed_date"] = t
                item["history"].append({"date": t, "event": "closed"})
                stats["closed"].append(item)

    for finding in remaining.values():
        item = make_item(next_id(meta), finding, sla, image_ref)
        poam["items"].append(item)
        stats["new"].append(item)

    # Apply / refresh waivers across all open items
    for item in poam["items"]:
        if item["status"] != "Open":
            item.pop("waiver", None)
            continue
        w = match_waiver(item, waivers)
        if w:
            item["waiver"] = {
                "reason": w.get("reason", ""),
                "expires": w.get("expires", ""),
                "approved_by": w.get("approved_by", ""),
            }
        else:
            item.pop("waiver", None)

    return stats


def gate_check(stats: dict, fail_on: str) -> list:
    fail_on = norm_severity(fail_on) if fail_on.upper() != "NONE" else "NONE"
    if fail_on == "NONE":
        return []
    threshold = SEVERITY_RANK[fail_on]
    failing = []
    for item in stats["new"] + stats["reopened"]:
        if SEVERITY_RANK[item["severity"]] >= threshold and "waiver" not in item:
            failing.append(item)
    return failing


# --------------------------------------------------------------------------
# XLSX report
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _severity_cell(ws, row, col, sev):
    cell = ws.cell(row=row, column=col, value=sev)
    cell.fill = PatternFill("solid", fgColor=SEVERITY_FILL.get(sev, "95A5A6"))
    cell.font = Font(bold=True, color="FFFFFF" if sev in ("CRITICAL", "HIGH", "UNKNOWN") else "000000")
    return cell


def days_between(start: str, end: str) -> int:
    try:
        return (date.fromisoformat(end) - date.fromisoformat(start)).days
    except (ValueError, TypeError):
        return 0


def write_xlsx(poam: dict, stats: dict, failing: list, gate_passed: bool, out_path: Path):
    meta = poam["metadata"]
    open_items = sorted(
        (i for i in poam["items"] if i["status"] == "Open"),
        key=lambda i: (-SEVERITY_RANK[i["severity"]], i["poam_id"]),
    )
    closed_items = sorted(
        (i for i in poam["items"] if i["status"] == "Closed"),
        key=lambda i: (i.get("closed_date") or "", i["poam_id"]),
        reverse=True,
    )

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Plan of Action & Milestones (POA&M)"
    ws["A1"].font = Font(bold=True, size=16)
    rows = [
        ("System", meta.get("system_name", "")),
        ("Image", meta.get("image_ref", "")),
        ("POA&M created", meta.get("created", "")),
        ("Last updated", meta.get("last_updated", "")),
        ("Scan number", meta.get("scan_count", 0)),
        ("Gate result", "PASS" if gate_passed else "FAIL"),
        ("", ""),
        ("New this run", len(stats["new"])),
        ("Reopened this run", len(stats["reopened"])),
        ("Closed this run", len(stats["closed"])),
        ("Gate-failing findings", len(failing)),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=2, value=value)
        if label == "Gate result":
            c = ws.cell(row=r, column=2)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="27AE60" if gate_passed else "C0392B")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Severity").font = BOLD
    ws.cell(row=r, column=2, value="Open").font = BOLD
    ws.cell(row=r, column=3, value="Closed").font = BOLD
    ws.cell(row=r, column=4, value="Total").font = BOLD
    r += 1
    for sev in SEVERITY_ORDER:
        o = sum(1 for i in open_items if i["severity"] == sev)
        c = sum(1 for i in closed_items if i["severity"] == sev)
        if o == 0 and c == 0 and sev == "UNKNOWN":
            continue
        _severity_cell(ws, r, 1, sev)
        ws.cell(row=r, column=2, value=o)
        ws.cell(row=r, column=3, value=c)
        ws.cell(row=r, column=4, value=o + c)
        r += 1
    ws.cell(row=r, column=1, value="Total").font = BOLD
    ws.cell(row=r, column=2, value=len(open_items)).font = BOLD
    ws.cell(row=r, column=3, value=len(closed_items)).font = BOLD
    ws.cell(row=r, column=4, value=len(open_items) + len(closed_items)).font = BOLD
    _set_widths(ws, [26, 44, 10, 10])

    # ---- Open sheet ----
    ws = wb.create_sheet("Open")
    headers = [
        "POA&M ID", "Severity", "CVE", "Package", "Installed", "Fixed In",
        "First Detected", "Last Seen", "Due Date", "Days Open", "Overdue",
        "Waived", "Title", "Targets", "Image",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    t = today_iso()
    for r, item in enumerate(open_items, start=2):
        overdue = item["scheduled_completion"] < t
        ws.cell(row=r, column=1, value=item["poam_id"])
        _severity_cell(ws, r, 2, item["severity"])
        ws.cell(row=r, column=3, value=item["cve_id"])
        ws.cell(row=r, column=4, value=item["package"])
        ws.cell(row=r, column=5, value=item["installed_version"])
        ws.cell(row=r, column=6, value=item["fixed_version"])
        ws.cell(row=r, column=7, value=item["first_detected"])
        ws.cell(row=r, column=8, value=item["last_seen"])
        ws.cell(row=r, column=9, value=item["scheduled_completion"])
        ws.cell(row=r, column=10, value=days_between(item["first_detected"], t))
        c = ws.cell(row=r, column=11, value="YES" if overdue else "")
        if overdue:
            c.font = Font(bold=True, color="C0392B")
        ws.cell(row=r, column=12, value="YES" if "waiver" in item else "")
        ws.cell(row=r, column=13, value=item["title"])
        ws.cell(row=r, column=14, value="; ".join(item["targets"]))
        ws.cell(row=r, column=15, value=item["image"])
    _set_widths(ws, [11, 11, 20, 26, 20, 20, 14, 14, 12, 10, 9, 8, 60, 40, 36])
    if open_items:
        ws.auto_filter.ref = ws.dimensions

    # ---- Closed sheet ----
    ws = wb.create_sheet("Closed")
    headers = [
        "POA&M ID", "Severity", "CVE", "Package", "First Detected",
        "Closed Date", "Days to Close", "Title", "Image",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r, item in enumerate(closed_items, start=2):
        ws.cell(row=r, column=1, value=item["poam_id"])
        _severity_cell(ws, r, 2, item["severity"])
        ws.cell(row=r, column=3, value=item["cve_id"])
        ws.cell(row=r, column=4, value=item["package"])
        ws.cell(row=r, column=5, value=item["first_detected"])
        ws.cell(row=r, column=6, value=item.get("closed_date") or "")
        ws.cell(row=r, column=7, value=days_between(item["first_detected"], item.get("closed_date") or today_iso()))
        ws.cell(row=r, column=8, value=item["title"])
        ws.cell(row=r, column=9, value=item["image"])
    _set_widths(ws, [11, 11, 20, 26, 14, 12, 13, 60, 36])
    if closed_items:
        ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)


# --------------------------------------------------------------------------
# GitHub integration
# --------------------------------------------------------------------------

def write_github_outputs(values: dict):
    out = os.environ.get("GITHUB_OUTPUT")
    if not out:
        return
    with open(out, "a") as f:
        for k, v in values.items():
            f.write(f"{k}={v}\n")


def write_step_summary(poam, stats, failing, gate_passed, mode):
    path = os.environ.get("GITHUB_STEP_SUMMARY")
    if not path:
        return
    meta = poam["metadata"]
    open_count = sum(1 for i in poam["items"] if i["status"] == "Open")
    closed_count = sum(1 for i in poam["items"] if i["status"] == "Closed")
    icon = "✅" if gate_passed else "❌"
    lines = [
        f"## {icon} POA&M Security Gate — {'PASS' if gate_passed else 'FAIL'}",
        "",
        f"**System:** {meta.get('system_name','')}  |  **Image:** `{meta.get('image_ref','')}`  |  **Mode:** {mode}  |  **Scan #:** {meta.get('scan_count')}",
        "",
        "| Metric | Count |",
        "| --- | ---: |",
        f"| New this run | {len(stats['new'])} |",
        f"| Reopened this run | {len(stats['reopened'])} |",
        f"| Closed this run | {len(stats['closed'])} |",
        f"| Total open | {open_count} |",
        f"| Total closed | {closed_count} |",
    ]
    if failing:
        lines += ["", "### Gate-failing findings", "", "| POA&M ID | Severity | CVE | Package | Fixed In |", "| --- | --- | --- | --- | --- |"]
        for item in failing[:20]:
            lines.append(
                f"| {item['poam_id']} | {item['severity']} | {item['cve_id']} | {item['package']} | {item['fixed_version'] or '—'} |"
            )
        if len(failing) > 20:
            lines.append(f"| … | | +{len(failing) - 20} more | | |")
    with open(path, "a") as f:
        f.write("\n".join(lines) + "\n")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="POA&M generator/validator for Trivy results")
    ap.add_argument("--trivy-results", required=True,
                    help="Comma-separated path(s) to Trivy JSON result files")
    ap.add_argument("--poam-file", default="",
                    help="Path to the previous run's POA&M JSON; missing/empty => initial mode")
    ap.add_argument("--output-dir", default="poam-out")
    ap.add_argument("--system-name", required=True)
    ap.add_argument("--image-ref", default="")
    ap.add_argument("--fail-on", default="CRITICAL",
                    help="Minimum severity of NEW/REOPENED findings that fails the gate (NONE disables)")
    ap.add_argument("--sla", default="",
                    help="Override remediation SLAs, e.g. CRITICAL=15,HIGH=30,MEDIUM=90,LOW=180")
    ap.add_argument("--waivers-file", default="",
                    help="Optional JSON list of risk-accepted CVEs with expiry dates")
    args = ap.parse_args()

    try:
        sla = parse_sla(args.sla)
        findings = load_trivy([p.strip() for p in args.trivy_results.split(",") if p.strip()])
        waivers = load_waivers(args.waivers_file)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return EXIT_ERROR

    poam_path = Path(args.poam_file) if args.poam_file else None
    if poam_path and poam_path.is_file():
        mode = "validate"
        poam = json.loads(poam_path.read_text())
    else:
        mode = "initial"
        poam = new_poam(args.system_name, args.image_ref)

    stats = process(poam, findings, sla, args.image_ref, waivers)
    failing = gate_check(stats, args.fail_on)
    gate_passed = not failing

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "poam.json"
    xlsx_path = out_dir / "poam.xlsx"
    json_path.write_text(json.dumps(poam, indent=2) + "\n")
    write_xlsx(poam, stats, failing, gate_passed, xlsx_path)

    open_count = sum(1 for i in poam["items"] if i["status"] == "Open")
    closed_count = sum(1 for i in poam["items"] if i["status"] == "Closed")

    write_github_outputs({
        "mode": mode,
        "gate-result": "pass" if gate_passed else "fail",
        "open-count": open_count,
        "closed-count": closed_count,
        "new-count": len(stats["new"]),
        "reopened-count": len(stats["reopened"]),
        "closed-this-run": len(stats["closed"]),
        "failing-count": len(failing),
        "poam-json": str(json_path),
        "poam-xlsx": str(xlsx_path),
    })
    write_step_summary(poam, stats, failing, gate_passed, mode)

    print(f"POA&M mode:          {mode}")
    print(f"New findings:        {len(stats['new'])}")
    print(f"Reopened findings:   {len(stats['reopened'])}")
    print(f"Closed this run:     {len(stats['closed'])}")
    print(f"Total open / closed: {open_count} / {closed_count}")
    print(f"Outputs:             {json_path}, {xlsx_path}")

    if not gate_passed:
        print(f"\nGATE FAILED: {len(failing)} new/reopened finding(s) at or above "
              f"'{args.fail_on.upper()}' severity:", file=sys.stderr)
        for item in failing:
            print(f"  {item['poam_id']}  {item['severity']:<8} {item['cve_id']}  "
                  f"{item['package']} ({item['installed_version']})", file=sys.stderr)
        return EXIT_GATE_FAILED

    print("\nGATE PASSED")
    return EXIT_OK


if __name__ == "__main__":
    sys.exit(main())
